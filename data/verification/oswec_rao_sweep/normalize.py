#!/usr/bin/env python3
"""
Normalize OSWEC RAO WEC-Sim reference data.

Raw input:
  raw/RAO_dat.xlsx  (sheet PTO_damping=0.0)
    Columns: Wave Period (s), Pitch, RAO
    RAO = Pitch / wave_height (0.01 m) in WEC-Sim convention.
    Multiply by 2 to convert to amplitude-based RAO (rad/m).

Output (normalized/):
  wecsim/pitch_rao.txt — omega (rad/s), pitch RAO (rad/m).
"""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import numpy as np
from normalize_utils import write_rao

HERE = os.path.dirname(os.path.abspath(__file__))


def main():
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required (pip install openpyxl)")
        sys.exit(1)

    xlsx = os.path.join(HERE, 'raw', 'RAO_dat.xlsx')
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb['PTO_damping=0.0']

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    periods, raos = [], []
    for row in rows:
        if row[0] is None:
            break
        periods.append(float(row[0]))
        rao_idx = header.index('RAO')
        raos.append(float(row[rao_idx]))

    omega = 2.0 * np.pi / np.array(periods)
    rao = np.array(raos) * 2.0  # WEC-Sim xlsx uses wave-height convention; convert to amplitude-based RAO

    data = np.column_stack([omega, rao])
    outpath = os.path.join(HERE, 'normalized', 'wecsim', 'pitch_rao.txt')
    write_rao(
        outpath, data,
        source='wecsim', model='oswec_rao_sweep',
        quantity='pitch_rao',
        units='rad/s, rad/m',
        columns='Omega(rad/s) PitchRAO(rad/m)',
    )
    print("Wrote normalized WEC-Sim OSWEC pitch RAO file")


if __name__ == '__main__':
    main()
