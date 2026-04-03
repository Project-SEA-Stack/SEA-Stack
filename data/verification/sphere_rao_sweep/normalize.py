#!/usr/bin/env python3
"""
Normalize sphere RAO multicode reference data.

Raw input:
  raw/RAO_dat.xlsx  (sheet PTO_S=0.002)
    Columns: Wave Period (s), InWave-HOTINT, InWave (Lin), Marin (NLin),
             NREL (NLin), ProteusDS (Lin)
    Values are heave RAO in m/m.

Output (normalized/):
  One file per source: omega (rad/s), heave RAO amplitude (m/m).
"""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import numpy as np
from normalize_utils import write_rao

HERE = os.path.dirname(os.path.abspath(__file__))

SOURCES = {
    'InWave-HOTINT': 'inwave_hotint',
    'InWave (Lin)':  'inwave',
    'Marin (NLin)':  'marin',
    'NREL (NLin)':   'nrel',
    'ProteusDS (Lin)': 'proteusds',
}


def main():
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required (pip install openpyxl)")
        sys.exit(1)

    xlsx = os.path.join(HERE, 'raw', 'RAO_dat.xlsx')
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb['PTO_S=0.002']

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    periods, source_data = [], {v: [] for v in SOURCES.values()}
    for row in rows:
        if row[0] is None:
            break
        periods.append(float(row[0]))
        for col_name, dir_name in SOURCES.items():
            col_idx = header.index(col_name)
            source_data[dir_name].append(float(row[col_idx]))

    omega = 2.0 * np.pi / np.array(periods)

    for dir_name, rao_vals in source_data.items():
        rao = np.array(rao_vals)
        data = np.column_stack([omega, rao])
        outpath = os.path.join(HERE, 'normalized', dir_name, 'heave_rao.txt')
        write_rao(
            outpath, data,
            source=dir_name, model='sphere_rao_sweep',
            quantity='heave_rao',
            units='rad/s, m/m',
            columns='Omega(rad/s) HeaveRAO(m/m)',
        )

    print(f"Wrote {len(SOURCES)} normalized sphere RAO files")


if __name__ == '__main__':
    main()
